from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path

import openpyxl
from tqdm import tqdm

from openrouter_api import OpenRouterAPI

# ── Configuration ─────────────────────────────────────────────────────────────

MODE = "RATIONALE"  # "RATIONALE" or "STRUCTURED"

API_KEY_PATH = Path("/home/ked/sync/tokens/sonata_openrouter.txt")
REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = REPO_ROOT / "parameters.xlsx"
OUTPUTS_DIR = REPO_ROOT / "outputs"

PROMPT_FILES = {
    "RATIONALE":   REPO_ROOT / "data" / "prompt_decision_simple.md",
    "STRUCTURED":  REPO_ROOT / "data" / "prompt_decision_struct.md",
}

# Column indices (1-based) for each mode
# Shared input columns
COL_TARGET_MODEL  = 3   # C
COL_EVALUATED_POST = 5  # E

# RATIONALE MODE output columns (F–K)
RATIONALE_COLS = {
    "model_id":       6,
    "prompt_id":      7,
    "generated_on":   8,
    "decision":       9,
    "short_rationale": 10,
    "reviewer_notes": 11,
}

# STRUCTURED OUTPUT MODE output columns (L–U)
STRUCTURED_COLS = {
    "model_id":                12,
    "prompt_id":               13,
    "generated_on":            14,
    "decision":                15,
    "evidence_spans":          16,
    "factors_for_decision":    17,
    "factors_against_decision": 18,
    "counterfactual_change":   19,
    "short_rationale":         20,
    "reviewer_notes":          21,
}

OUTPUT_COLS = {
    "RATIONALE":  RATIONALE_COLS,
    "STRUCTURED": STRUCTURED_COLS,
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_json_response(text: str) -> dict:
    # Strip markdown code fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text.strip(), flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text.strip())
    return json.loads(text.strip())


def join_list(value: object) -> str:
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    return str(value) if value is not None else ""


def build_row_data(parsed: dict, mode: str, target_model: str, prompt_id: str) -> dict[str, object]:
    today = date.today().isoformat()
    base = {
        "model_id":     target_model,
        "prompt_id":    prompt_id,
        "generated_on": today,
        "decision":     parsed.get("decision", ""),
        "reviewer_notes": "",
    }
    if mode == "RATIONALE":
        base["short_rationale"] = parsed.get("rationale", "")
    else:
        base["evidence_spans"]           = join_list(parsed.get("evidence_spans"))
        base["factors_for_decision"]     = join_list(parsed.get("factors_for_decision"))
        base["factors_against_decision"] = join_list(parsed.get("factors_against_decision"))
        base["counterfactual_change"]    = parsed.get("counterfactual_change", "")
        base["short_rationale"]          = parsed.get("short_rationale", "")
    return base


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    cols = OUTPUT_COLS[MODE]
    prompt_path = PROMPT_FILES[MODE]
    prompt_template = prompt_path.read_text(encoding="utf-8")
    prompt_id = prompt_path.name

    OUTPUTS_DIR.mkdir(exist_ok=True)
    raw_outputs_path = OUTPUTS_DIR / f"eval_outputs_{MODE.lower()}_{date.today().isoformat()}.json"

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["eval"]

    # Collect rows to process
    data_rows = []
    for row_idx in range(3, ws.max_row + 1):
        post = ws.cell(row_idx, COL_EVALUATED_POST).value
        model = ws.cell(row_idx, cols["model_id"]).value
        if not post:
            continue
        if model:  # already filled for this mode — skip
            continue
        data_rows.append(row_idx)

    if not data_rows:
        print(f"Nothing to evaluate (mode={MODE}). All rows already filled or no posts found.")
        return

    # Cache API clients by model name
    clients: dict[str, OpenRouterAPI] = {}

    def get_client(model_name: str) -> OpenRouterAPI:
        if model_name not in clients:
            clients[model_name] = OpenRouterAPI(
                model_name=model_name,
                api_key_path=str(API_KEY_PATH),
            )
        return clients[model_name]

    MAX_ATTEMPTS = 3
    raw_outputs: dict[str, dict] = {}

    with tqdm(total=len(data_rows), desc=f"Evaluating [{MODE}]", unit="post") as progress:
        for row_idx in data_rows:
            post_text = str(ws.cell(row_idx, COL_EVALUATED_POST).value).strip()
            target_model = str(ws.cell(row_idx, COL_TARGET_MODEL).value or "").strip()
            pair_id = ws.cell(row_idx, 1).value
            case_side = ws.cell(row_idx, 2).value

            if not target_model:
                tqdm.write(f"[{pair_id}/{case_side}] SKIP: target_model is empty")
                progress.update(1)
                continue

            prompt = prompt_template.replace("{post}", post_text)
            client = get_client(target_model)

            parsed = None
            last_err: Exception | None = None
            for attempt in range(1, MAX_ATTEMPTS + 1):
                try:
                    raw = client.generate(prompt)
                    parsed = parse_json_response(raw)
                    last_err = None
                    break
                except Exception as exc:
                    last_err = exc
                    if attempt < MAX_ATTEMPTS:
                        tqdm.write(f"[{pair_id}/{case_side}] attempt {attempt} failed ({exc}) — retrying…")

            if last_err or parsed is None:
                tqdm.write(f"[{pair_id}/{case_side}] ERROR: {last_err}")
                progress.update(1)
                continue

            raw_outputs[f"{pair_id}/{case_side}"] = parsed

            row_data = build_row_data(parsed, MODE, target_model, prompt_id)
            for field, col_idx in cols.items():
                ws.cell(row_idx, col_idx).value = row_data.get(field, "")

            progress.set_postfix(pair=f"{pair_id}{case_side}", decision=parsed.get("decision", "?"))
            progress.update(1)

    wb.save(XLSX_PATH)
    raw_outputs_path.write_text(
        json.dumps(raw_outputs, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Saved results to {XLSX_PATH.name}")
    print(f"Saved raw outputs to {raw_outputs_path.name}")


if __name__ == "__main__":
    main()
